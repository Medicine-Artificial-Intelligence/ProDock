import os
import time
import argparse
import pandas as pd
import shutil
import prolif as plf
import MDAnalysis as mda
import tempfile
from rdkit import Chem
from rdkit.Chem import Draw
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter as get_excel_column_letter
from PIL import Image as PILImage
import glob
import matplotlib

matplotlib.use("Agg")


def run_admetlab_evaluation_with_ketcher(Smiles, output_dir, compound_id, max_retries=1):
    """
    Submits a SMILES string to ADMETLab 3.0 (with the Ketcher widget),
    downloads CSV and PDF results, and draws a radar plot.
    Special characters in the SMILES (e.g. (), =, @) are passed safely
    via Selenium’s JS argument API to avoid any quoting issues.
    """
    downloaded_exts = set()
    target_names = {"csv": f"{compound_id}_admetlab_results.csv", "pdf": f"{compound_id}_admetlab_summary.pdf"}

    # 1) Check for already-downloaded files
    for ext, fname in target_names.items():
        if os.path.exists(os.path.join(output_dir, fname)):
            downloaded_exts.add(ext)

    # 2) Try up to max_retries times
    for attempt in range(1, max_retries + 1):
        if downloaded_exts == {"csv", "pdf"}:
            print(f"[✓] All files already downloaded for {compound_id}. Skipping.")
            break

        print(f"[{compound_id}] Attempt {attempt} (missing: {', '.join({'csv','pdf'} - downloaded_exts)})")
        # Setup headless Chrome with download prefs
        chrome_options = Options()
        prefs = {
            "download.default_directory": os.path.abspath(output_dir),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
            "profile.default_content_setting_values.automatic_downloads": 1,
        }
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_argument("--headless")  # enable if desired

        driver = webdriver.Chrome(options=chrome_options)
        wait = WebDriverWait(driver, 15)

        try:
            # 2a) Load the evaluation page
            driver.get("https://admetlab3.scbdd.com/server/evaluation")
            wait.until(lambda d: d.execute_script("return document.readyState") == "complete")

            # 2b) Dump all input/textarea IDs (debug)
            for elt in driver.find_elements(By.CSS_SELECTOR, "input, textarea"):
                print(
                    elt.tag_name,
                    "id=",
                    elt.get_attribute("id"),
                    "name=",
                    elt.get_attribute("name"),
                    "placeholder=",
                    elt.get_attribute("placeholder"),
                )

            # 2c) Locate the SMILES input by its real ID (replace "smiles" if different)
            input_box = driver.find_element(By.ID, "smiles")

            # 2d) Clean up and strip
            Smiles_clean = Smiles.strip()
            print(f"[→] Injecting SMILES exactly as in CSV: {Smiles_clean!r}")

            # 2e) Clear, then send via JS argument to handle special chars
            input_box.clear()
            time.sleep(0.1)
            driver.execute_script(
                "arguments[0].value = arguments[1]; "
                "arguments[0].dispatchEvent(new Event('input',{bubbles:true})); "
                "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                input_box,
                Smiles_clean,
            )
            time.sleep(0.1)

            # 2f) Verify it landed
            current = input_box.get_attribute("value")
            print(f"[→] SMILES now in box: {current!r}")

            # 2g) Submit the form
            submit_button = driver.find_element(By.XPATH, '//button[contains(text(), "Submit")]')
            submit_button.click()
            print(f"[✓] Clicked submit for {compound_id}")

            # 2h) Wait for CSV/PDF download buttons
            wait.until(EC.element_to_be_clickable((By.ID, "download_csv")))
            wait.until(EC.element_to_be_clickable((By.ID, "download_pdf")))

            # 2i) Trigger downloads
            if "csv" not in downloaded_exts:
                driver.find_element(By.ID, "download_csv").click()
                print(f"[✓] CSV download triggered for {compound_id}")
            if "pdf" not in downloaded_exts:
                driver.find_element(By.ID, "download_pdf").click()
                print(f"[✓] PDF download triggered for {compound_id}")

            # 2j) Wait & rename downloaded files
            for ext in ["csv", "pdf"]:
                if ext in downloaded_exts:
                    continue
                success = False
                for _ in range(20):
                    files = [f for f in os.listdir(output_dir) if f.endswith(f".{ext}")]
                    if files:
                        path = rename_latest_file(output_dir, ext, target_names[ext])
                        if path:
                            downloaded_exts.add(ext)
                            success = True
                        break
                    time.sleep(0.1)
                if not success:
                    print(f"[!] Missing .{ext} after attempt {attempt} for {compound_id}")

            # 2k) Generate radar plot if both are present
            if downloaded_exts == {"csv", "pdf"}:
                csv_path = os.path.join(output_dir, target_names["csv"])
                df = pd.read_csv(csv_path)
                if not df.empty:
                    props = df.iloc[0].to_dict()
                    radar_path = os.path.join(output_dir, f"{compound_id}_radar.png")
                    draw_radar_plot(compound_id, props, radar_path)
                break

        except Exception as e:
            print(f"[✗] Error in attempt {attempt} for {compound_id}: {e}")
        finally:
            driver.quit()
    else:
        print(f"[X] Failed to process {compound_id} after {max_retries} attempts.")


def rename_latest_file(folder, extension, new_name):
    files = [f for f in os.listdir(folder) if f.endswith(f".{extension}")]
    if not files:
        print(f"[!] No .{extension} file found in {folder}")
        return None

    files = sorted([os.path.join(folder, f) for f in files], key=os.path.getmtime, reverse=True)
    latest = files[0]
    new_path = os.path.join(folder, new_name)

    if os.path.abspath(latest) == os.path.abspath(new_path):
        print(f"[✓] File already named: {new_name}")
        return latest

    try:
        shutil.move(latest, new_path)
        print(f"[✓] Renamed {os.path.basename(latest)} → {new_name}")
        return new_path
    except Exception as e:
        print(f"[✗] Failed to rename {latest} to {new_name}: {e}")
        return None


def draw_molecule(Smiles, path):
    mol = Chem.MolFromSmiles(Smiles)
    if mol:
        img = Draw.MolToImage(mol, size=(300, 300))
        img.save(path)
    else:
        print(f"[!] Invalid Smiles: {Smiles}")


def draw_radar_plot(compound_name, props, save_path):
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors
    import numpy as np

    keys = ["MW", "logP", "logS", "logD", "nHA", "nHD", "TPSA", "nRot", "nRing", "MaxRing", "nHet", "fChar", "nRig"]

    upper_limit = [600, 3, 0.5, 3, 12, 7, 140, 11, 6, 18, 15, 4, 30]
    lower_limit = [100, 0, -4, 1, 0, 0, 0, 0, 0, 0, 1, -4, 0]

    buffer_percent = [0.5, 0.10, 0.40, 0.05, 0.5, 0.25, 0.55, 0.30, 0.65, 0.15, 0.35, 0.35, 0.05]

    values = [props.get(k, 0) for k in keys]

    plot_min = []
    plot_max = []
    for i, val in enumerate(values):
        lo = min(lower_limit[i], val)
        hi = max(upper_limit[i], val)
        buf = buffer_percent[i]
        range_ = hi - lo if hi != lo else 1.0
        plot_min.append(lo - range_ * buf)
        plot_max.append(hi + range_ * buf)

    def normalize(val, lo, hi):
        return (val - lo) / (hi - lo) if hi != lo else 0.5

    norm_values = [normalize(v, lo, hi) for v, lo, hi in zip(values, plot_min, plot_max)]
    norm_upper = [normalize(u, lo, hi) for u, lo, hi in zip(upper_limit, plot_min, plot_max)]
    norm_lower = [normalize(lim, lo, hi) for lim, lo, hi in zip(lower_limit, plot_min, plot_max)]

    norm_values += norm_values[:1]
    norm_upper += norm_upper[:1]
    norm_lower += norm_lower[:1]
    angles = np.linspace(0, 2 * np.pi, len(keys), endpoint=False).tolist()
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))

    ax.plot(angles, norm_upper, color="blue", linewidth=1.5, label="Upper Limit")
    ax.fill(angles, norm_upper, color="blue", alpha=0.1)
    ax.plot(angles, norm_lower, color="green", linewidth=1.5, label="Lower Limit")
    ax.fill(angles, norm_lower, color="green", alpha=0.1)

    cmap = mcolors.LinearSegmentedColormap.from_list("compound_grad", ["yellow", "orange", "yellow"])
    norm = plt.Normalize(0, len(keys) - 1)

    for i in range(len(keys)):
        ax.plot(angles[i : i + 2], norm_values[i : i + 2], color=cmap(norm(i)), linewidth=2)

    for i, key in enumerate(keys):
        real = values[i]
        if real < lower_limit[i] or real > upper_limit[i]:
            ax.plot(angles[i], norm_values[i], "o", color="red", markersize=8)
        else:
            ax.plot(angles[i], norm_values[i], "o", color=cmap(norm(i)), markersize=8)

    ax.set_thetagrids(np.degrees(angles[:-1]), keys)
    ax.set_ylim(0, 1)
    ax.set_yticklabels([])
    plt.tight_layout()
    plt.savefig(save_path)
    plt.close()


def screenshot_lignetwork(protein_path, ligand_sdf, output_png):
    """
    Generate a lignetwork HTML for the first pose in ligand_sdf,
    run the fingerprint analysis (using all available CPUs), render it via Selenium, and save a PNG.
    """
    # 1) Load protein and create ProLIF Molecule
    u = mda.Universe(protein_path)
    protein_mol = plf.Molecule.from_mda(u)

    # 2) Read all poses from the SDF and ensure at least one
    poses = list(plf.sdf_supplier(ligand_sdf))
    if not poses:
        raise ValueError(f"No molecules found in SDF file {ligand_sdf}")
    first_pose = poses[0]

    # 3) Run the fingerprint analysis with cpu=-1 (use all cores)
    fp = plf.Fingerprint().run_from_iterable(
        poses,
        protein_mol,
    )

    # 4) Plot the lignetwork for the first pose
    html_obj = fp.plot_lignetwork(first_pose)

    # 5) Write the HTML to a temporary file
    html_str = html_obj.data
    with tempfile.NamedTemporaryFile(suffix=".html", delete=False, mode="w") as f:
        f.write(html_str)
        html_path = f.name

    # 6) Launch headless Chrome, load the HTML, and take a screenshot
    options = Options()
    options.headless = True
    options.add_argument("--window-size=1000,900")
    driver = webdriver.Chrome(options=options)
    driver.get(f"file://{html_path}")
    time.sleep(2)  # allow any JS to finish rendering
    driver.save_screenshot(output_png)
    driver.quit()


def merge_and_draw(output_dir):
    merged_csv = os.path.join(output_dir, "merged_admet_results.csv")
    merged_xlsx = os.path.join(output_dir, "merged_admet_results.xlsx")
    compound_data = []
    image_info = {}
    has_cluster = False

    # Collect per-compound data and track images
    for subdir in os.listdir(output_dir):
        subpath = os.path.join(output_dir, subdir)
        if not os.path.isdir(subpath):
            continue

        for file in os.listdir(subpath):
            if file.endswith(".csv") and "_admetlab_results" in file:
                df = pd.read_csv(os.path.join(subpath, file))

                if "smiles" in df.columns:
                    df.rename(columns={"smiles": "Smiles"}, inplace=True)

                # Determine compound and optional cluster from folder name
                # If folder contains "_", assume final segment may be cluster only
                # when it was created from an input Cluster column.
                compound = subdir
                cluster = None

                metadata_path = os.path.join(subpath, "metadata.json")
                if os.path.exists(metadata_path):
                    with open(metadata_path, "r") as f:
                        metadata = json.load(f)
                    compound = metadata.get("Compound", compound)
                    cluster = metadata.get("Cluster", None)

                df["Compound"] = compound

                if cluster is not None:
                    df["Cluster"] = cluster
                    has_cluster = True

                struct_img = os.path.join(subpath, f"{compound}.png")
                radar_img = os.path.join(subpath, f"{compound}_radar.png")
                plot2d_img = os.path.join(subpath, f"{compound}_2d_plot.png")

                image_info[len(compound_data)] = (struct_img, radar_img, plot2d_img)
                compound_data.append(df)

    if not compound_data:
        print("[!] No ADMET result CSVs found.")
        return

    # Merge and cleanup
    merged_df = pd.concat(compound_data, ignore_index=True)

    if "taskId" in merged_df.columns:
        merged_df.drop(columns=["taskId"], inplace=True)

    if "smiles" in merged_df.columns:
        merged_df.rename(columns={"smiles": "Smiles"}, inplace=True)

    # If Cluster was never provided, make sure it is not included
    if not has_cluster and "Cluster" in merged_df.columns:
        merged_df.drop(columns=["Cluster"], inplace=True)

    # Save merged CSV with preferred column ordering
    cols = merged_df.columns.tolist()
    ordered = []

    preferred = ["Compound"]
    if has_cluster:
        preferred.append("Cluster")
    preferred.append("Smiles")

    for key in preferred:
        match = next((c for c in cols if c.lower() == key.lower()), None)
        if match:
            ordered.append(match)
            cols.remove(match)

    ordered += cols
    merged_df[ordered].to_csv(merged_csv, index=False)
    print(f"[✓] Merged results saved: {merged_csv}")

    # Prepare Excel with images
    xlsx_df = merged_df.copy()
    xlsx_cols = xlsx_df.columns.tolist()
    ordered_xlsx = []

    for key in preferred:
        match = next((c for c in xlsx_cols if c.lower() == key.lower()), None)
        if match:
            ordered_xlsx.append(match)
            xlsx_cols.remove(match)

    smiles_col = next((c for c in ordered_xlsx if c.lower() == "smiles"), None)
    if not smiles_col:
        raise ValueError("'Smiles' column not found.")

    pos = ordered_xlsx.index(smiles_col) + 1
    for col_name in ["structure", "radar_plot", "2d_plot"]:
        ordered_xlsx.insert(pos, col_name)
        pos += 1

    ordered_xlsx += [c for c in xlsx_cols if c not in ordered_xlsx]
    xlsx_df = xlsx_df.reindex(columns=ordered_xlsx)
    xlsx_df.to_excel(merged_xlsx, index=False)

    wb = load_workbook(merged_xlsx)
    ws = wb.active
    letter_map = {col: get_excel_column_letter(i + 1) for i, col in enumerate(xlsx_df.columns)}

    # Embed images and adjust sizes
    width_2d = None

    for idx, (_, row) in enumerate(xlsx_df.iterrows()):
        row_num = idx + 2
        struct_img, radar_img, plot2d_img = image_info.get(idx, ("", "", ""))

        for colname, img_path in [
            ("structure", struct_img),
            ("radar_plot", radar_img),
            ("2d_plot", plot2d_img),
        ]:
            if img_path and os.path.exists(img_path):
                img = XLImage(img_path)

                if colname == "2d_plot":
                    orig = PILImage.open(img_path)
                    w, h = orig.size
                    img.height = 100
                    img.width = int(w / h * 100)
                    width_2d = img.width
                else:
                    img.width = img.height = 100

                letter = letter_map.get(colname)
                if letter:
                    ws.add_image(img, f"{letter}{row_num}")

        ws.row_dimensions[row_num].height = 80

    # Adjust column widths
    for col in ["structure", "radar_plot", "Compound"]:
        letter = letter_map.get(col)
        if letter:
            ws.column_dimensions[letter].width = 20

    if width_2d is not None:
        letter = letter_map.get("2d_plot")
        if letter:
            ws.column_dimensions[letter].width = max(int(width_2d / 7), 20)

    letter_sm = letter_map.get("Smiles")
    if letter_sm:
        ws.column_dimensions[letter_sm].width = 80

    wb.save(merged_xlsx)
    print(f"[✓] Excel file with embedded images saved: {merged_xlsx}")


def main(args):
    # ensure output directory exists
    os.makedirs(args.output_dir, exist_ok=True)

    # validate: if a protein is given, GNINA outputs must be too
    if args.protein_path and not args.gnina_output_dir:
        raise ValueError("When --protein_path is specified, --gnina_output_dir must also be provided")

    # read compounds list
    df = pd.read_csv(args.input_csv)

    required_cols = {"Compounds", "Smiles"}
    missing_cols = required_cols - set(df.columns)
    if missing_cols:
        raise ValueError(f"Input CSV is missing required columns: {', '.join(missing_cols)}")

    has_cluster = "Cluster" in df.columns

    protein_path = args.protein_path
    gnina_output_dir = args.gnina_output_dir

    for _, row in df.iterrows():
        compound = str(row["Compounds"])
        Smiles = row["Smiles"]

        if has_cluster and pd.notna(row["Cluster"]):
            cluster = str(row["Cluster"])
            folder_name = f"{compound}_{cluster}"
        else:
            cluster = None
            folder_name = compound

        folder = os.path.join(args.output_dir, folder_name)
        os.makedirs(folder, exist_ok=True)

        # Save metadata so merge_and_draw does not have to guess
        metadata = {"Compound": compound}
        if cluster is not None:
            metadata["Cluster"] = cluster

        with open(os.path.join(folder, "metadata.json"), "w") as f:
            json.dump(metadata, f, indent=2)

        # 2D structure
        draw_molecule(Smiles, os.path.join(folder, f"{compound}.png"))

        # ADMETLab evaluation
        run_admetlab_evaluation_with_ketcher(Smiles, folder, compound)

        # optional 2D lignetwork plot via GNINA output
        if protein_path:
            rank = row.get("Best_satisfied_rank_gnina", row.get("Best_satisfied_rank"))

            sdf_path = os.path.join(gnina_output_dir, str(rank), f"{compound}.sdf")

            if not os.path.exists(sdf_path):
                pattern = os.path.join(gnina_output_dir, "*", str(rank), f"{compound}.sdf")
                matches = glob.glob(pattern)

                if matches:
                    sdf_path = matches[0]
                else:
                    print(f"[!] SDF not found for {compound} at rank {rank}")
                    continue

            screenshot_lignetwork(protein_path, sdf_path, os.path.join(folder, f"{compound}_2d_plot.png"))

    merge_and_draw(args.output_dir)
    print("[✔] All compounds processed and results merged.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Batch ADMETLab 3.0 evaluator with optional 2D lignetwork plots")
    parser.add_argument(
        "--input_csv", required=True, help="CSV with columns: Compounds, Smiles. Optional column: Cluster"
    )
    parser.add_argument("--output_dir", required=True, help="Base output directory for results")
    parser.add_argument(
        "--protein_path", required=False, help="Path to the single .pdb protein file (triggers 2D plot when used)"
    )
    parser.add_argument(
        "--gnina_output_dir",
        required=False,
        help="Directory containing GNINA outputs; subfolders are ranks (required if --protein_path is used)",
    )
    args = parser.parse_args()
    main(args)
